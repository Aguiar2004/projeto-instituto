from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os
from threading import Lock

app = Flask(__name__)
app.secret_key = "troque-essa-chave"

ARQUIVO_EXCEL = "cadastros.xlsx"
excel_lock = Lock()


def criar_planilha_se_nao_existir():
    if not os.path.exists(ARQUIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cadastros"

        headers = [
            "Data do cadastro",
            "Tipo de cadastro",
            "Nome completo do usuário",
            "Data de nascimento",
            "CPF do usuário",
            "Nome completo do responsável legal",
            "Grau de parentesco",
            "CPF do responsável legal",
            "Telefone/WhatsApp",
            "Endereço",
            "Bairro/Comunidade",
            "Sexo",
            "Outro sexo",
            "Gênero",
            "Outro gênero",
            "Idade",
            "Raça/Cor",
            "Escolaridade",
            "Frequenta escola atualmente",
            "Escola matriculada",
            "Estado civil",
            "Mora com companheiro(a)",
            "Tem filhos",
            "Quantidade de filhos",
            "Pessoas na residência",
            "Pessoas com renda",
            "Situação de moradia",
            "Outro tipo de moradia",
            "Tempo de moradia",
            "Ocupação",
            "Situação atual",
            "Outro situação atual",
            "Renda familiar mensal",
            "Recebe benefício",
            "Bolsa Família",
            "BPC",
            "Outro benefício",
            "Como conheceu o Instituto",
            "Outro meio de conhecimento",
            "Já participou de alguma atividade",
            "Observações técnicas"
        ]

        ws.append(headers)
        wb.save(ARQUIVO_EXCEL)


@app.route("/")
def index():
    return redirect(url_for("formulario"))


@app.route("/formulario", methods=["GET", "POST"])
def formulario():
    if request.method == "POST":
        dados = {
            "data_cadastro": request.form.get("data_cadastro", "").strip(),
            "tipo_cadastro": request.form.get("tipo_cadastro", "").strip(),
            "nome_completo": request.form.get("nome_completo", "").strip(),
            "data_nascimento": request.form.get("data_nascimento", "").strip(),
            "cpf_usuario": request.form.get("cpf_usuario", "").strip(),
            "nome_responsavel": request.form.get("nome_responsavel", "").strip(),
            "grau_parentesco": request.form.get("grau_parentesco", "").strip(),
            "cpf_responsavel": request.form.get("cpf_responsavel", "").strip(),
            "telefone": request.form.get("telefone", "").strip(),
            "endereco": request.form.get("endereco", "").strip(),
            "bairro": request.form.get("bairro", "").strip(),
            "sexo": request.form.get("sexo", "").strip(),
            "outro_sexo": request.form.get("outro_sexo", "").strip(),
            "genero": request.form.get("genero", "").strip(),
            "outro_genero": request.form.get("outro_genero", "").strip(),
            "idade": request.form.get("idade", "").strip(),
            "raca_cor": request.form.get("raca_cor", "").strip(),
            "escolaridade": request.form.get("escolaridade", "").strip(),
            "frequenta_escola": request.form.get("frequenta_escola", "").strip(),
            "escola_matriculada": request.form.get("escola_matriculada", "").strip(),
            "estado_civil": request.form.get("estado_civil", "").strip(),
            "mora_companheiro": request.form.get("mora_companheiro", "").strip(),
            "tem_filhos": request.form.get("tem_filhos", "").strip(),
            "quantidade_filhos": request.form.get("quantidade_filhos", "").strip(),
            "pessoas_residencia": request.form.get("pessoas_residencia", "").strip(),
            "pessoas_renda": request.form.get("pessoas_renda", "").strip(),
            "situacao_moradia": request.form.get("situacao_moradia", "").strip(),
            "outro_moradia": request.form.get("outro_moradia", "").strip(),
            "tempo_moradia": request.form.get("tempo_moradia", "").strip(),
            "ocupacao": request.form.get("ocupacao", "").strip(),
            "situacao_atual": request.form.get("situacao_atual", "").strip(),
            "outro_situacao_atual": request.form.get("outro_situacao_atual", "").strip(),
            "renda_familiar": request.form.get("renda_familiar", "").strip(),
            "recebe_beneficio": request.form.get("recebe_beneficio", "").strip(),
            "bolsa_familia": "Sim" if request.form.get("bolsa_familia") == "on" else "Não",
            "bpc": "Sim" if request.form.get("bpc") == "on" else "Não",
            "outro_beneficio": request.form.get("outro_beneficio", "").strip(),
            "como_conheceu": request.form.get("como_conheceu", "").strip(),
            "outro_conheceu": request.form.get("outro_conheceu", "").strip(),
            "ja_participou": request.form.get("ja_participou", "").strip(),
            "observacoes": request.form.get("observacoes", "").strip(),
        }

        if not dados["nome_completo"]:
            flash("O campo Nome completo é obrigatório.", "erro")
            return redirect(url_for("formulario"))

        try:
            with excel_lock:
                wb = load_workbook(ARQUIVO_EXCEL)
                ws = wb["Cadastros"]

                ws.append([
                    dados["data_cadastro"],
                    dados["tipo_cadastro"],
                    dados["nome_completo"],
                    dados["data_nascimento"],
                    dados["cpf_usuario"],
                    dados["nome_responsavel"],
                    dados["grau_parentesco"],
                    dados["cpf_responsavel"],
                    dados["telefone"],
                    dados["endereco"],
                    dados["bairro"],
                    dados["sexo"],
                    dados["outro_sexo"],
                    dados["genero"],
                    dados["outro_genero"],
                    dados["idade"],
                    dados["raca_cor"],
                    dados["escolaridade"],
                    dados["frequenta_escola"],
                    dados["escola_matriculada"],
                    dados["estado_civil"],
                    dados["mora_companheiro"],
                    dados["tem_filhos"],
                    dados["quantidade_filhos"],
                    dados["pessoas_residencia"],
                    dados["pessoas_renda"],
                    dados["situacao_moradia"],
                    dados["outro_moradia"],
                    dados["tempo_moradia"],
                    dados["ocupacao"],
                    dados["situacao_atual"],
                    dados["outro_situacao_atual"],
                    dados["renda_familiar"],
                    dados["recebe_beneficio"],
                    dados["bolsa_familia"],
                    dados["bpc"],
                    dados["outro_beneficio"],
                    dados["como_conheceu"],
                    dados["outro_conheceu"],
                    dados["ja_participou"],
                    dados["observacoes"]
                ])

                wb.save(ARQUIVO_EXCEL)

            flash("Cadastro salvo com sucesso.", "sucesso")
            return redirect(url_for("formulario"))

        except Exception as e:
            flash(f"Erro ao salvar cadastro: {str(e)}", "erro")
            return redirect(url_for("formulario"))

    return render_template("formulario.html")


if __name__ == "__main__":
    criar_planilha_se_nao_existir()
    app.run(debug=True)